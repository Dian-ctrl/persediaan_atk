#!/usr/bin/env python3
"""
generate_laporan.py - Laporan Persediaan BPJS Ketenagakerjaan
Format persis sama dengan laporan_persediaan__per_31012026.xlsx
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def thin():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def medium_bottom():
    t = Side(style='thin'); m = Side(style='medium')
    return Border(left=t, right=t, top=t, bottom=m)

def cell_style(cell, val=None, bold=False, sz=9, h='center', v='center',
               wrap=False, rgb=None, brd=None, fmt=None, italic=False, underline=False):
    if val is not None:
        cell.value = val
    u = 'single' if underline else None
    cell.font = Font(name='Arial', size=sz, bold=bold, italic=italic, underline=u)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if rgb:
        cell.fill = PatternFill('solid', fgColor=rgb)
    if brd:
        cell.border = brd
    if fmt:
        cell.number_format = fmt

def fill_range(ws, r1, c1, r2, c2, rgb, brd=None):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill('solid', fgColor=rgb)
            if brd:
                cell.border = brd

def hdr_merge(ws, r1, c1, r2, c2, txt, rgb='BDD7EE'):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(r1, c1)
    cell_style(c, txt, bold=True, sz=9, h='center', v='center', wrap=True, rgb=rgb, brd=thin())
    fill_range(ws, r1, c1, r2, c2, rgb, thin())

# Kolom: A=1…O=15 (sama dengan asli)
# A=No, B=Nama, C=Satuan
# D=Saldo_awal_qty
# E=Saldo_awal_harga, F=Saldo_awal_nilai
# G=Masuk_qty, H=Masuk_harga, I=Masuk_nilai
# J=Keluar_qty, K=Keluar_harga, L=Keluar_nilai
# M=Akhir_qty, N=Akhir_harga, O=Akhir_nilai

def build(data):
    wb = Workbook()
    ws = wb.active
    ws.title = data['bulan_indo'][:3].upper() + str(data['tahun'])[2:]

    # ── Lebar kolom (persis asli) ──
    widths = {1:6.3, 2:50, 3:10.5, 4:9, 5:14, 6:18,
              7:9, 8:14, 9:22, 10:9, 11:14, 12:18, 13:9, 14:20, 15:18}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── ROW 1: kosong ──
    ws.row_dimensions[1].height = 12

    # ── ROW 2: Judul ──
    ws.merge_cells('A2:O2')
    cell_style(ws['A2'],
        'Daftar Stock Opname Alat Tulis Kantor, Barang Cetakan, Meterai dan Consumable',
        bold=True, sz=12, h='center')
    ws.row_dimensions[2].height = 20

    # ── ROW 3: Periode ──
    ws.merge_cells('A3:O3')
    cell_style(ws['A3'],
        f"Per  {data['hari']} {data['bulan_indo']} {data['tahun']}",
        bold=True, sz=11, h='center')
    ws.row_dimensions[3].height = 18

    # ── ROW 4: spasi ──
    ws.row_dimensions[4].height = 6

    # ── ROW 5-6: Header tabel ──
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 42
    B = 'BDD7EE'

    # Kolom tetap (merge baris 5 dan 6)
    hdr_merge(ws,5,1,6,1,'No')
    hdr_merge(ws,5,2,6,2,'Jenis Barang')
    hdr_merge(ws,5,3,6,3,'Satuan')
    hdr_merge(ws,5,4,6,4,'Saldo')

    # Saldo awal
    bulan_sebelum_label = f"Saldo {data['hari_akhir_sebelum']} {data['bulan_sebelum_indo']} {data['tahun_sebelum']}"
    hdr_merge(ws,5,5,5,6, bulan_sebelum_label)
    cell_style(ws.cell(6,5),'Harga Satuan',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,6),'Total Harga (Rp)',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())

    # Penambahan
    hdr_merge(ws,5,7,5,9, f"Penambahan {data['bulan_indo']} {data['tahun']}")
    cell_style(ws.cell(6,7),'Saldo',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,8),'Harga Satuan',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,9),'Total Harga (Rp)',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())

    # Pemakaian
    hdr_merge(ws,5,10,5,12, f"Pemakaian {data['bulan_indo']} {data['tahun']}")
    cell_style(ws.cell(6,10),'Saldo',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,11),'Harga Satuan',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,12),'Total Harga (Rp)',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())

    # Saldo akhir
    hdr_merge(ws,5,13,5,15, f"Saldo {data['hari']} {data['bulan_indo']} {data['tahun']}")
    cell_style(ws.cell(6,13),'Saldo',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,14),'Harga Satuan',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())
    cell_style(ws.cell(6,15),'Total Harga (Rp)',bold=True,sz=9,h='center',v='center',wrap=True,rgb=B,brd=thin())

    # ── DATA ROWS ──
    row = 7
    IDR = '#,##0'
    CAT_Y = 'FFF2CC'  # kuning kategori
    TOT_G = 'E2EFDA'  # hijau subtotal
    GT_GR = 'D9D9D9'  # abu grand total

    for kat in data['kategoris']:
        # ── Baris kategori ──
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=15)
        cell_style(ws.cell(row,1), kat['label'], bold=True, sz=9,
                   h='left', v='center', rgb=CAT_Y, brd=thin())
        fill_range(ws, row,1,row,15, CAT_Y, thin())
        row += 1

        # ── Baris barang ──
        for no, item in enumerate(kat['items']):
            ws.row_dimensions[row].height = 15

            def d(col, val, h='center', fmt=None):
                c = ws.cell(row, col)
                display = val if (val is not None and val != 0) else None
                cell_style(c, display, sz=9, h=h, v='center', brd=thin(), fmt=fmt)

            d(1, no+1)
            d(2, item['nama'], h='left')
            d(3, item['satuan'])
            d(4, item['saldo_qty'] or None)
            d(5, item['saldo_harga'] if item['saldo_harga'] > 0 else None, fmt=IDR)
            d(6, item['saldo_nilai'] or None, fmt=IDR)
            d(7, item['masuk_qty'] or None)
            d(8, item['masuk_harga'] if item['masuk_qty'] > 0 else None, fmt=IDR)
            d(9, item['masuk_nilai'] or None, fmt=IDR)
            d(10, item['keluar_qty'] or None)
            d(11, item['keluar_harga'] if item['keluar_qty'] > 0 else None, fmt=IDR)
            d(12, item['keluar_nilai'] or None, fmt=IDR)
            d(13, item['akhir_qty'] or None)
            d(14, item['akhir_harga'] if item['akhir_qty'] > 0 else None, fmt=IDR)
            d(15, item['akhir_nilai'] or None, fmt=IDR)
            row += 1

        # ── Baris kosong ──
        ws.row_dimensions[row].height = 5
        for c in range(1,16): ws.cell(row,c).border = thin()
        row += 1

        # ── Sub total kategori ──
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
        s = kat['subtotal']
        label_total = 'Total ' + kat['label'].replace('/ 0','/ ').split('/')[0].strip().title()
        cell_style(ws.cell(row,1), label_total, bold=True, sz=9,
                   h='center', v='center', rgb=TOT_G, brd=thin())
        fill_range(ws,row,1,row,3,TOT_G,thin())

        def st(col, val, fmt=IDR):
            c = ws.cell(row, col)
            cell_style(c, val if val != 0 else None, bold=True, sz=9,
                       h='center', v='center', rgb=TOT_G, brd=thin(), fmt=fmt)

        st(4, s['saldo_qty'], '#,##0')
        st(5, None)
        st(6, s['saldo_nilai'])
        st(7, s['masuk_qty'], '#,##0')
        st(8, None)
        st(9, s['masuk_nilai'])
        st(10, s['keluar_qty'], '#,##0')
        st(11, None)
        st(12, s['keluar_nilai'])
        st(13, s['akhir_qty'], '#,##0')
        st(14, None)
        st(15, s['akhir_nilai'])
        row += 1

        # ── Spasi antar kategori ──
        ws.row_dimensions[row].height = 5
        row += 1

    # ── GRAND TOTAL ──
    ws.row_dimensions[row].height = 18
    g = data['grand_total']
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
    cell_style(ws.cell(row,1),'Grand Total',bold=True,sz=10,
               h='center',v='center',rgb=GT_GR,brd=thin())
    fill_range(ws,row,1,row,3,GT_GR,thin())

    def gt(col, val, fmt=IDR):
        c = ws.cell(row, col)
        cell_style(c, val if val != 0 else None, bold=True, sz=10,
                   h='center', v='center', rgb=GT_GR, brd=thin(), fmt=fmt)

    gt(4, g['saldo_qty'], '#,##0')
    gt(5, None)
    gt(6, g['saldo_nilai'])
    gt(7, g['masuk_qty'], '#,##0')
    gt(8, None)
    gt(9, g['masuk_nilai'])
    gt(10, g['keluar_qty'], '#,##0')
    gt(11, None)
    gt(12, g['keluar_nilai'])
    gt(13, g['akhir_qty'], '#,##0')
    gt(14, None)
    gt(15, g['akhir_nilai'])

    # ── TANDA TANGAN ──
    row += 3
    pg = data['pengaturan']

    # Tanggal kanan atas
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    cell_style(ws.cell(row,13),
               f"{data['kota']}, {data['hari']} {data['bulan_indo']} {data['tahun']}",
               sz=9, h='right')

    row += 2
    # Label: Mengetahui / Menyetujui / Membuat
    for col_s, lbl in [(3,'Mengetahui,'), (9,'Menyetujui,'), (14,'Membuat,')]:
        ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_s+1)
        cell_style(ws.cell(row,col_s), lbl, sz=9, h='center')

    # Spasi tanda tangan (5 baris)
    row += 5

    # Nama dan jabatan
    for col_s, nm_key, jb_key in [
        (3, 'nama_mengetahui', 'jabatan_mengetahui'),
        (9, 'nama_menyetujui', 'jabatan_menyetujui'),
        (14,'nama_membuat',   'jabatan_membuat'),
    ]:
        nm = pg.get(nm_key, '')
        jb = pg.get(jb_key, '')
        ws.merge_cells(start_row=row, start_column=col_s, end_row=row, end_column=col_s+2)
        cell_style(ws.cell(row,col_s), nm, bold=True, sz=9, h='center', underline=True)
        ws.merge_cells(start_row=row+1, start_column=col_s, end_row=row+1, end_column=col_s+2)
        cell_style(ws.cell(row+1,col_s), jb, sz=9, h='center')

    # ── Page setup (landscape A3, mirip asli) ──
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A3
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left=0.4; ws.page_margins.right=0.4
    ws.page_margins.top=0.6;  ws.page_margins.bottom=0.6
    ws.print_title_rows = '5:6'
    ws.sheet_view.zoomScale = 75

    return wb


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: generate_laporan.py <data.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding='utf-8') as f:
        data = json.load(f)
    wb = build(data)
    wb.save(sys.argv[2])
    print(f"OK: {sys.argv[2]}")
